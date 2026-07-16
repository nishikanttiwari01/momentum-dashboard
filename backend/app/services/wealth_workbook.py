from __future__ import annotations

from datetime import date, datetime
from hashlib import sha256
from io import BytesIO
import math
import re
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.datetime import to_excel
from pydantic import BaseModel, ConfigDict, computed_field

from app.schemas.wealth_portfolio import IGNORED_SHEETS, ImportIssue


SUPPORTED_SHEETS = frozenset({
    "BALANCE SHEET",
    "CURRENT ASSET",
    "FUNDS",
    "Funds XIRR",
    "Final XIRR",
    "EQUITY",
    "FIXED ASSET",
    "GOALS",
    "MNTHLY INCOM PLAN",
    "Gera office roi",
})


class ParsedAsset(BaseModel):
    source_key: str
    name: str
    asset_type: str
    category: str | None = None
    account_owner: str | None = None
    account_type: str | None = None
    market: str = "IN"
    currency: str = "INR"
    invested_amount: float | None = None
    market_value: float | None = None
    source_ref: dict[str, Any]


class ParsedTransaction(BaseModel):
    source_key: str
    asset_source_key: str
    occurred_on: date
    kind: str = "BUY"
    amount: float
    units: float | None = None
    unit_price: float | None = None
    currency: str = "INR"
    source_ref: dict[str, Any]


class ParsedValuation(BaseModel):
    source_key: str
    asset_source_key: str
    valued_on: date
    market_value: float
    currency: str = "INR"
    source_ref: dict[str, Any]


class HouseholdAggregate(BaseModel):
    model_config = ConfigDict(frozen=True)

    as_of_label: str
    financial_invested_capital: float | None
    financial_market_value: float | None
    property_invested_capital: float | None
    property_market_value: float | None
    invested_capital: float | None
    total_market_value: float | None
    monthly_rent: float | None


class FixedAssetComponent(BaseModel):
    model_config = ConfigDict(frozen=True)

    name: str
    description: str | None = None
    invested_amount: float | None = None
    market_value: float
    source_ref: dict[str, Any]


class ParsedLedgerAsset(BaseModel):
    model_config = ConfigDict(frozen=True)

    source_key: str
    owner: str | None = None
    name: str
    category: str | None = None
    asset_class: str
    market: str = "IN"
    currency: str = "INR"
    source_ref: dict[str, Any]


class ParsedAssetObservation(BaseModel):
    model_config = ConfigDict(frozen=True)

    source_key: str
    asset_source_key: str
    observed_on: date
    principal: float | None = None
    market_value: float | None = None
    currency: str = "INR"
    source_ref: dict[str, Any]


class ParsedLedgerCashFlow(BaseModel):
    model_config = ConfigDict(frozen=True)

    source_key: str
    asset_source_key: str | None = None
    occurred_on: date
    flow_type: str
    amount: float
    currency: str = "INR"
    source_ref: dict[str, Any]


class ParsedReportingSource(BaseModel):
    model_config = ConfigDict(frozen=True)

    metric: str
    source_sheet: str
    source_cell: str
    observed_on: date | None = None


class ParsedReportingPeriod(BaseModel):
    model_config = ConfigDict(frozen=True)

    year: int
    label: str
    sources: tuple[ParsedReportingSource, ...]
    controls: dict[str, float | None] = {}


class ParsedWorkbook(BaseModel):
    source_sha256: str
    filename: str
    recognized_sheets: list[str]
    ignored_sheets: list[str]
    assets: list[ParsedAsset]
    transactions: list[ParsedTransaction]
    valuations: list[ParsedValuation]
    household: HouseholdAggregate | None = None
    fixed_assets: tuple[FixedAssetComponent, ...] = ()
    ledger_assets: tuple[ParsedLedgerAsset, ...] = ()
    asset_observations: tuple[ParsedAssetObservation, ...] = ()
    ledger_cash_flows: tuple[ParsedLedgerCashFlow, ...] = ()
    reporting_periods: tuple[ParsedReportingPeriod, ...] = ()
    reconciliation_warnings: tuple[ImportIssue, ...] = ()
    issues: list[ImportIssue]

    @computed_field
    @property
    def counts(self) -> dict[str, int]:
        return {
            "assets": len(self.assets),
            "transactions": len(self.transactions),
            "valuations": len(self.valuations),
        }


def _header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).casefold()


def _name(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def _number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (date, datetime)):
        number = float(to_excel(value))
        return number if math.isfinite(number) else None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def _date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        text = value.strip().splitlines()[0]
        for pattern in ("%Y-%m-%d", "%d %b %Y", "%d-%m-%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(text, pattern).date()
            except ValueError:
                continue
    return None


def _cell_ref(sheet: str, row: int, column: int) -> dict[str, Any]:
    from openpyxl.utils import get_column_letter
    return {"sheet": sheet, "row": row, "column": column, "cell": f"{get_column_letter(column)}{row}"}


def _parse_current_asset_ledger(sheet) -> tuple[list[ParsedLedgerAsset], list[ParsedAssetObservation]]:
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 3:
        return [], []
    dates, roles = rows[0], rows[1]
    assets: list[ParsedLedgerAsset] = []
    observations: list[ParsedAssetObservation] = []
    for row_number, row in enumerate(rows[2:], start=3):
        owner = _name(row[0] if row else None)
        name = _name(row[1] if len(row) > 1 else None)
        if not owner or not name or owner.casefold() in {"sub total", "subtotal", "total", "notes"}:
            continue
        category = _name(row[2] if len(row) > 2 else None) or None
        asset_key = _source_key("ledger", "CURRENT ASSET", owner, name, category)
        assets.append(ParsedLedgerAsset(
            source_key=asset_key, owner=owner, name=name, category=category,
            asset_class="financial", source_ref=_cell_ref("CURRENT ASSET", row_number, 2),
        ))
        column = 4
        while column <= len(roles):
            role = _header(roles[column - 1] if column - 1 < len(roles) else None)
            observed_on = _date(dates[column - 1] if column - 1 < len(dates) else None)
            if observed_on is None:
                column += 1
                continue
            if "principal" in role:
                next_role = _header(roles[column] if column < len(roles) else None)
                market = _number(row[column] if column < len(row) and "mkt value" in next_role else None)
                principal = _number(row[column - 1] if column - 1 < len(row) else None)
                if principal is not None or market is not None:
                    observations.append(ParsedAssetObservation(
                        source_key=_source_key("observation", asset_key, observed_on, principal, market),
                        asset_source_key=asset_key, observed_on=observed_on, principal=principal,
                        market_value=market, source_ref=_cell_ref("CURRENT ASSET", row_number, column),
                    ))
                column += 2 if "mkt value" in next_role else 1
                continue
            if "mkt value" in role:
                market = _number(row[column - 1] if column - 1 < len(row) else None)
                if market is not None:
                    observations.append(ParsedAssetObservation(
                        source_key=_source_key("observation", asset_key, observed_on, None, market),
                        asset_source_key=asset_key, observed_on=observed_on, market_value=market,
                        source_ref=_cell_ref("CURRENT ASSET", row_number, column),
                    ))
            column += 1
    return assets, observations


def _parse_fixed_asset_ledger(sheet) -> tuple[list[ParsedLedgerAsset], list[ParsedAssetObservation], list[ParsedLedgerCashFlow]]:
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return [], [], []
    assets: list[ParsedLedgerAsset] = []
    observations: list[ParsedAssetObservation] = []
    assets_by_name: dict[str, ParsedLedgerAsset] = {}
    headings = rows[0]
    for row_number, row in enumerate(rows[1:], start=2):
        if not isinstance(row[0] if row else None, (int, float)):
            continue
        name = _name(row[1] if len(row) > 1 else None)
        if not name:
            continue
        category = _name(row[2] if len(row) > 2 else None) or None
        asset_key = _source_key("ledger", "FIXED ASSET", name, category)
        asset = ParsedLedgerAsset(
            source_key=asset_key, name=name, category=category, asset_class="property",
            source_ref=_cell_ref("FIXED ASSET", row_number, 2),
        )
        assets.append(asset)
        assets_by_name[name.casefold()] = asset
        column = 4
        while column <= len(headings):
            observed_on = _date(headings[column - 1])
            role = _header(headings[column - 1])
            if observed_on is not None and "principal" in role:
                principal = _number(row[column - 1] if column - 1 < len(row) else None)
                next_role = _header(headings[column] if column < len(headings) else None)
                market = _number(row[column] if column < len(row) and "mkt value" in next_role else None)
                if principal is not None or market is not None:
                    observations.append(ParsedAssetObservation(
                        source_key=_source_key("observation", asset_key, observed_on, principal, market),
                        asset_source_key=asset_key, observed_on=observed_on, principal=principal,
                        market_value=market, source_ref=_cell_ref("FIXED ASSET", row_number, column),
                    ))
                column += 2
            else:
                column += 1
    flows: list[ParsedLedgerCashFlow] = []
    brigade = next((asset for name, asset in assets_by_name.items() if "brigade" in name), None)
    for heading_index, row in enumerate(rows):
        date_columns = [index for index, value in enumerate(row) if _header(value) == "date"]
        if not date_columns:
            continue
        for row_number, values in enumerate(rows[heading_index + 1:], start=heading_index + 2):
            if any(_header(value).startswith("total") for value in values if isinstance(value, str)):
                break
            for date_index in date_columns:
                occurred_on = _date(values[date_index] if date_index < len(values) else None)
                amount = _number(values[date_index + 1] if date_index + 1 < len(values) else None)
                if occurred_on is not None and amount is not None:
                    flows.append(ParsedLedgerCashFlow(
                        source_key=_source_key("property_capital", brigade.source_key if brigade else None, occurred_on, amount, date_index),
                        asset_source_key=brigade.source_key if brigade else None, occurred_on=occurred_on,
                        flow_type="property_capital", amount=amount,
                        source_ref=_cell_ref("FIXED ASSET", row_number, date_index + 1),
                    ))
        break
    return assets, observations, flows


_REPORT_METRICS = {
    "current assests principal (year end)": "financial_principal",
    "current assets market value (year end)": "financial_market_value",
    "fixed assests principal (year end)": "property_principal",
    "fixed assests market value (year end)": "property_market_value",
}


def _resolve_direct_reference(source_sheets, sheet_name: str, cell_ref: str) -> tuple[str, str]:
    for _ in range(8):
        sheet = source_sheets.get(sheet_name)
        if sheet is None:
            break
        value = sheet[cell_ref].value
        if not isinstance(value, str):
            break
        match = re.fullmatch(r"='?([^']+)'?!\$?([A-Z]+\$?\d+)", value.strip())
        if match:
            sheet_name, cell_ref = match.group(1), match.group(2).replace("$", "")
            continue
        local_match = re.fullmatch(r"=\$?([A-Z]+\$?\d+)", value.strip())
        if local_match:
            cell_ref = local_match.group(1).replace("$", "")
            continue
        break
    return sheet_name, cell_ref


def _parse_reporting_periods(formula_sheet, value_sheet, source_sheets) -> list[ParsedReportingPeriod]:
    formula_rows = list(formula_sheet.iter_rows(values_only=True))
    value_rows = list(value_sheet.iter_rows(values_only=True))
    header_index = next((i for i, row in enumerate(formula_rows) if _header(row[0] if row else None) == "asset type"), None)
    if header_index is None:
        return []
    periods: list[ParsedReportingPeriod] = []
    for column in range(1, len(formula_rows[header_index])):
        label = _name(formula_rows[header_index][column])
        year_match = re.search(r"(\d{4})", label)
        if not year_match:
            continue
        sources: list[ParsedReportingSource] = []
        controls: dict[str, float | None] = {}
        for row_index in range(header_index + 1, len(formula_rows)):
            label_key = _header(formula_rows[row_index][0] if formula_rows[row_index] else None)
            value = formula_rows[row_index][column] if column < len(formula_rows[row_index]) else None
            cached = value_rows[row_index][column] if row_index < len(value_rows) and column < len(value_rows[row_index]) else None
            if label_key in _REPORT_METRICS and isinstance(value, str):
                match = re.fullmatch(r"='?([^']+)'?!\$?([A-Z]+\$?\d+)", value.strip())
                if match:
                    source_sheet, source_cell = match.groups()
                    source_sheet, source_cell = _resolve_direct_reference(
                        source_sheets, source_sheet, source_cell.replace("$", "")
                    )
                    source_column = re.match(r"[A-Z]+", source_cell.replace("$", "")).group()
                    source = source_sheets.get(source_sheet)
                    observed_on = _date(source[f"{source_column}1"].value) if source is not None else None
                    sources.append(ParsedReportingSource(
                        metric=_REPORT_METRICS[label_key], source_sheet=source_sheet,
                        source_cell=source_cell.replace("$", ""), observed_on=observed_on,
                    ))
            if label_key:
                controls[label_key] = _number(cached)
        periods.append(ParsedReportingPeriod(
            year=int(year_match.group(1)), label=label, sources=tuple(sources), controls=controls,
        ))
    return periods


def _source_key(*parts: Any) -> str:
    normalized = "|".join(_name(part).casefold() for part in parts)
    return sha256(normalized.encode("utf-8")).hexdigest()


def _unique_source_items(items):
    return list({item.source_key: item for item in items}.values())


def _rows(sheet, required_heading: str | None = None) -> tuple[dict[str, int], list[tuple[int, tuple[Any, ...]]]]:
    all_rows = list(sheet.iter_rows(values_only=True))
    header_index = 0
    if required_heading:
        for index, row in enumerate(all_rows[:20]):
            if required_heading.casefold() in {_header(value) for value in row}:
                header_index = index
                break
    if not all_rows:
        return {}, []
    headings = all_rows[header_index]
    columns = {_header(value): index for index, value in enumerate(headings) if _header(value)}
    return columns, [
        (row_number, row)
        for row_number, row in enumerate(all_rows[header_index + 1 :], start=header_index + 2)
    ]


def _value(row: tuple[Any, ...], columns: dict[str, int], *names: str) -> Any:
    for name in names:
        index = columns.get(name.casefold())
        if index is not None and index < len(row):
            return row[index]
    return None


def _parse_funds(sheet) -> tuple[list[ParsedAsset], list[ImportIssue]]:
    columns, rows = _rows(sheet, "fund name")
    if "fund name" not in columns:
        columns, rows = _rows(sheet, "fund")
    assets_by_key: dict[str, ParsedAsset] = {}
    issues: list[ImportIssue] = []
    for row_number, row in rows:
        serial = _value(row, columns, "s no.", "s no", "sno")
        name = _name(_value(row, columns, "fund", "fund name", "scheme"))
        if not name:
            continue
        if "fund name" in columns and not isinstance(serial, (int, float)):
            continue
        invested = _number(_value(row, columns, "principal", "invested", "investment"))
        market_value = _number(_value(row, columns, "market value", "current value", "value"))
        if invested is None and market_value is None:
            issues.append(ImportIssue(
                severity="warning", code="fund_without_value",
                message=f"{name} has no principal or market value", sheet="FUNDS", row=row_number,
            ))
            continue
        asset = ParsedAsset(
            source_key=_source_key(
                "fund", _value(row, columns, "name", "owner"),
                _value(row, columns, "mode"), name, _value(row, columns, "acc type"),
            ),
            name=name,
            asset_type="mutual_fund",
            category=_name(_value(row, columns, "category")) or None,
            account_owner=_name(_value(row, columns, "name", "owner")) or None,
            account_type=_name(_value(row, columns, "acc type")) or None,
            invested_amount=invested,
            market_value=market_value,
            source_ref={"sheet": "FUNDS", "row": row_number},
        )
        existing = assets_by_key.get(asset.source_key)
        if existing is None:
            assets_by_key[asset.source_key] = asset
            continue

        source_rows = existing.source_ref.get("rows") or [existing.source_ref["row"]]
        assets_by_key[asset.source_key] = existing.model_copy(update={
            "invested_amount": sum(
                value for value in (existing.invested_amount, asset.invested_amount)
                if value is not None
            ),
            "market_value": sum(
                value for value in (existing.market_value, asset.market_value)
                if value is not None
            ),
            "source_ref": {"sheet": "FUNDS", "rows": [*source_rows, row_number]},
        })
    return list(assets_by_key.values()), issues


def _parse_flat_fund_transactions(sheet, assets: list[ParsedAsset]) -> tuple[list[ParsedTransaction], list[ParsedValuation], list[ImportIssue]]:
    columns, rows = _rows(sheet)
    transactions: list[ParsedTransaction] = []
    issues: list[ImportIssue] = []
    assets_by_name = {asset.name.casefold(): asset for asset in assets}
    for row_number, row in rows:
        name = _name(_value(row, columns, "fund", "fund name", "scheme"))
        if not name:
            continue
        occurred_on = _date(_value(row, columns, "date", "transaction date"))
        if occurred_on is None:
            issues.append(ImportIssue(
                severity="error", code="invalid_transaction_date",
                message=f"{name} has an invalid transaction date", sheet="Funds XIRR", row=row_number,
            ))
            continue
        amount = _number(_value(row, columns, "amount", "invested"))
        if amount is None or amount <= 0:
            issues.append(ImportIssue(
                severity="error", code="invalid_transaction_amount",
                message=f"{name} has an invalid transaction amount", sheet="Funds XIRR", row=row_number,
            ))
            continue
        asset = assets_by_name.get(name.casefold())
        if asset is None:
            issues.append(ImportIssue(
                severity="error", code="unknown_transaction_asset",
                message=f"{name} is not present in FUNDS", sheet="Funds XIRR", row=row_number,
            ))
            continue
        units = _number(_value(row, columns, "units"))
        nav = _number(_value(row, columns, "nav", "unit price"))
        transactions.append(ParsedTransaction(
            source_key=_source_key(
                "fund-buy", name, occurred_on.isoformat(), amount, units, nav,
                "row", row_number,
            ),
            asset_source_key=asset.source_key,
            occurred_on=occurred_on,
            amount=amount,
            units=units,
            unit_price=nav,
            source_ref={"sheet": "Funds XIRR", "row": row_number},
        ))
    return transactions, [], issues


def _parse_paired_fund_transactions(sheet, assets: list[ParsedAsset]) -> tuple[list[ParsedTransaction], list[ParsedValuation], list[ImportIssue]]:
    rows = list(sheet.iter_rows(values_only=True))
    transactions: list[ParsedTransaction] = []
    valuations: list[ParsedValuation] = []
    issues: list[ImportIssue] = []
    if len(rows) < 4:
        return transactions, valuations, issues

    assets_by_owner_name = {
        ((asset.account_owner or "").casefold(), asset.name.casefold()): asset
        for asset in assets
    }
    assets_by_name = {asset.name.casefold(): asset for asset in assets}
    width = max((len(row) for row in rows), default=0)
    for date_column in range(0, width, 2):
        amount_column = date_column + 1
        owner = _name(rows[0][date_column] if date_column < len(rows[0]) else None)
        fund_name = _name(rows[1][date_column] if date_column < len(rows[1]) else None)
        if not fund_name:
            continue
        asset = assets_by_owner_name.get((owner.casefold(), fund_name.casefold())) or assets_by_name.get(fund_name.casefold())
        if asset is None:
            issues.append(ImportIssue(
                severity="error", code="unknown_transaction_asset",
                message=f"{fund_name} is not present in FUNDS", sheet="Funds XIRR", row=2,
            ))
            continue
        for row_number, row in enumerate(rows[3:], start=4):
            raw_date = row[date_column] if date_column < len(row) else None
            raw_amount = row[amount_column] if amount_column < len(row) else None
            if raw_date is None and raw_amount is None:
                continue
            occurred_on = _date(raw_date)
            amount = _number(raw_amount)
            if (occurred_on is not None and raw_amount is None) or (raw_date is None and amount is not None):
                issues.append(ImportIssue(
                    severity="warning", code="incomplete_cash_flow",
                    message=f"{fund_name} has an incomplete cash-flow pair and was skipped",
                    sheet="Funds XIRR", row=row_number,
                ))
                continue
            if occurred_on is None or amount is None:
                issues.append(ImportIssue(
                    severity="error", code="invalid_transaction_date" if occurred_on is None else "invalid_transaction_amount",
                    message=f"{fund_name} has an invalid XIRR cash-flow row", sheet="Funds XIRR", row=row_number,
                ))
                continue
            if amount < 0:
                transactions.append(ParsedTransaction(
                    source_key=_source_key(
                        "fund-buy", owner, fund_name, occurred_on.isoformat(),
                        abs(amount), "row", row_number, "column", date_column + 1,
                    ),
                    asset_source_key=asset.source_key,
                    occurred_on=occurred_on,
                    amount=abs(amount),
                    source_ref={"sheet": "Funds XIRR", "row": row_number, "column": date_column + 1},
                ))
            elif amount > 0:
                valuations.append(ParsedValuation(
                    source_key=_source_key(
                        "valuation", owner, fund_name, occurred_on.isoformat(),
                        amount, "row", row_number, "column", date_column + 1,
                    ),
                    asset_source_key=asset.source_key,
                    valued_on=occurred_on,
                    market_value=amount,
                    source_ref={"sheet": "Funds XIRR", "row": row_number, "column": date_column + 1},
                ))
    return transactions, valuations, issues


def _parse_fund_transactions(sheet, assets: list[ParsedAsset]) -> tuple[list[ParsedTransaction], list[ParsedValuation], list[ImportIssue]]:
    first_row = next(sheet.iter_rows(values_only=True), ())
    normalized = {_header(value) for value in first_row}
    if "date" in normalized or "transaction date" in normalized:
        return _parse_flat_fund_transactions(sheet, assets)
    return _parse_paired_fund_transactions(sheet, assets)


def _parse_household(balance_sheet, income_sheet) -> tuple[HouseholdAggregate, list[ImportIssue]]:
    rows = list(balance_sheet.iter_rows(values_only=True))
    warnings: list[ImportIssue] = []
    header_row = next((row for row in rows if _header(row[0] if row else None) == "asset type"), ())
    populated_columns = [
        index for index in range(1, len(header_row))
        if _name(header_row[index]) and any(_number(row[index] if index < len(row) else None) is not None for row in rows)
    ]
    latest_column = max(populated_columns, default=1)
    as_of_label = _name(header_row[latest_column] if latest_column < len(header_row) else "") or "unknown"
    values: dict[str, float | None] = {}
    for row in rows:
        label = _header(row[0] if row else None)
        if label:
            values[label] = _number(row[latest_column] if latest_column < len(row) else None)

    financial_mv = values.get("current assets market value (year end)")
    property_mv = values.get("fixed assests market value (year end)")
    total_mv = values.get("total assets market value")
    financial_principal = values.get("current assests principal (year end)")
    property_principal = values.get("fixed assests principal (year end)")
    total_principal = values.get("total assets principal")

    rent_values: list[float] = []
    if income_sheet is not None:
        income_rows = list(income_sheet.iter_rows(values_only=True))
        heading = next((row for row in income_rows if "monthly income" in {_header(v) for v in row}), ())
        monthly_column = next((i for i, value in enumerate(heading) if _header(value) == "monthly income"), None)
        for row_number, row in enumerate(income_rows, 1):
            label = _header(row[0] if row else None)
            if "rent" not in label:
                continue
            rent = _number(row[monthly_column] if monthly_column is not None and monthly_column < len(row) else None)
            if rent is None:
                warnings.append(ImportIssue(severity="warning", code="missing_rent", message=f"{_name(row[0])} has no current monthly rent", sheet="MNTHLY INCOM PLAN", row=row_number))
            else:
                rent_values.append(rent)
    if income_sheet is None or not rent_values:
        warnings.append(ImportIssue(severity="warning", code="missing_rent", message="Current monthly rent is missing", sheet="MNTHLY INCOM PLAN"))
    if property_mv is None:
        warnings.append(ImportIssue(severity="warning", code="missing_property_value", message="Latest balance sheet property value is missing", sheet="BALANCE SHEET"))
    if financial_mv is not None and property_mv is not None and total_mv is not None and abs(financial_mv + property_mv - total_mv) > 1:
        warnings.append(ImportIssue(severity="warning", code="household_total_mismatch", message="Financial and property values do not reconcile to total assets within INR 1", sheet="BALANCE SHEET"))
    return HouseholdAggregate(
        as_of_label=as_of_label, financial_invested_capital=financial_principal,
        financial_market_value=financial_mv, property_invested_capital=property_principal,
        property_market_value=property_mv, invested_capital=total_principal,
        total_market_value=total_mv, monthly_rent=sum(rent_values) if rent_values else None,
    ), warnings


def _parse_fixed_assets(sheet) -> list[FixedAssetComponent]:
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headings = rows[0]
    market_columns = [i for i, value in enumerate(headings) if "mkt value" in _header(value)]
    market_column = max(market_columns, default=-1)
    principal_column = market_column - 1
    result: list[FixedAssetComponent] = []
    for row_number, row in enumerate(rows[1:], 2):
        name = _name(row[1] if len(row) > 1 else None)
        market_value = _number(row[market_column] if 0 <= market_column < len(row) else None)
        if not name or market_value is None:
            continue
        result.append(FixedAssetComponent(
            name=name, description=_name(row[2] if len(row) > 2 else None) or None,
            invested_amount=_number(row[principal_column] if principal_column < len(row) else None),
            market_value=market_value, source_ref={"sheet": "FIXED ASSET", "row": row_number},
        ))
    return result


def parse_workbook(payload: bytes, filename: str) -> ParsedWorkbook:
    source_sha256 = sha256(payload).hexdigest()
    workbook = load_workbook(BytesIO(payload), read_only=True, data_only=True)
    formula_workbook = load_workbook(BytesIO(payload), read_only=True, data_only=False)
    recognized = [name for name in workbook.sheetnames if name in SUPPORTED_SHEETS]
    ignored = [name for name in workbook.sheetnames if name in IGNORED_SHEETS]
    issues: list[ImportIssue] = []
    household = None
    fixed_assets: list[FixedAssetComponent] = []
    reconciliation_warnings: list[ImportIssue] = []
    ledger_assets: list[ParsedLedgerAsset] = []
    asset_observations: list[ParsedAssetObservation] = []
    ledger_cash_flows: list[ParsedLedgerCashFlow] = []
    reporting_periods: list[ParsedReportingPeriod] = []
    if "BALANCE SHEET" in recognized:
        household, reconciliation_warnings = _parse_household(
            workbook["BALANCE SHEET"], workbook["MNTHLY INCOM PLAN"] if "MNTHLY INCOM PLAN" in recognized else None,
        )
        issues.extend(reconciliation_warnings)
    if "FIXED ASSET" in recognized:
        fixed_assets = _parse_fixed_assets(workbook["FIXED ASSET"])
        property_assets, property_observations, ledger_cash_flows = _parse_fixed_asset_ledger(
            workbook["FIXED ASSET"]
        )
        ledger_assets.extend(property_assets)
        asset_observations.extend(property_observations)
    if "CURRENT ASSET" in recognized:
        financial_assets, financial_observations = _parse_current_asset_ledger(workbook["CURRENT ASSET"])
        ledger_assets.extend(financial_assets)
        asset_observations.extend(financial_observations)
    if "BALANCE SHEET" in recognized:
        reporting_periods = _parse_reporting_periods(
            formula_workbook["BALANCE SHEET"], workbook["BALANCE SHEET"],
            {name: formula_workbook[name] for name in ("CURRENT ASSET", "FIXED ASSET") if name in formula_workbook.sheetnames},
        )

    assets: list[ParsedAsset] = []
    if "FUNDS" in recognized:
        assets, fund_issues = _parse_funds(workbook["FUNDS"])
        issues.extend(fund_issues)

    transactions: list[ParsedTransaction] = []
    valuations: list[ParsedValuation] = []
    if "Funds XIRR" in recognized:
        transactions, valuations, transaction_issues = _parse_fund_transactions(workbook["Funds XIRR"], assets)
        issues.extend(transaction_issues)

    if not valuations:
        latest_transaction_date = max((item.occurred_on for item in transactions), default=date.today())
        valuations = [
            ParsedValuation(
                source_key=_source_key("valuation", asset.source_key, latest_transaction_date.isoformat(), asset.market_value),
                asset_source_key=asset.source_key,
                valued_on=latest_transaction_date,
                market_value=asset.market_value,
                source_ref=asset.source_ref,
            )
            for asset in assets
            if asset.market_value is not None
        ]

    parsed_sheets = {"FUNDS", "Funds XIRR", "BALANCE SHEET", "FIXED ASSET", "MNTHLY INCOM PLAN"}
    for sheet_name in recognized:
        if sheet_name not in parsed_sheets:
            issues.append(ImportIssue(
                severity="warning",
                code="sheet_parser_pending",
                message=f"{sheet_name} was recognized but is not normalized in the foundation phase",
            ))

    return ParsedWorkbook(
        source_sha256=source_sha256,
        filename=filename,
        recognized_sheets=recognized,
        ignored_sheets=ignored,
        assets=assets,
        transactions=transactions,
        valuations=valuations,
        household=household,
        fixed_assets=tuple(fixed_assets),
        ledger_assets=tuple(ledger_assets),
        asset_observations=tuple(_unique_source_items(asset_observations)),
        ledger_cash_flows=tuple(_unique_source_items(ledger_cash_flows)),
        reporting_periods=tuple(reporting_periods),
        reconciliation_warnings=tuple(reconciliation_warnings),
        issues=issues,
    )
