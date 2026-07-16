from io import BytesIO

from openpyxl import load_workbook

from tests.fixtures.wealth_workbook_factory import (
    make_real_layout_workbook_bytes,
    make_source_ledger_workbook_bytes,
    make_workbook_bytes,
)

from app.services.wealth_workbook import parse_workbook


def test_parser_extracts_assets_transactions_and_valuations():
    result = parse_workbook(make_workbook_bytes(), "investment.xlsx")
    assert result.counts == {"assets": 1, "transactions": 1, "valuations": 1}
    assert result.assets[0].name == "Example Mid Cap"
    assert result.transactions[0].occurred_on.isoformat() == "2025-03-06"


def test_parser_reconciles_latest_household_wealth_without_double_counting_fixed_components():
    result = parse_workbook(make_workbook_bytes(include_household=True), "investment.xlsx")
    assert result.household.as_of_label == "FY-2026"
    assert result.household.financial_market_value == 44_658_852.25
    assert result.household.property_market_value == 38_400_000
    assert result.household.total_market_value == 83_058_852.25
    assert result.household.invested_capital == 58_663_055.25
    assert result.household.monthly_rent == 44_000
    assert [item.market_value for item in result.fixed_assets] == [21_600_000, 6_800_000, 10_000_000]
    assert sum(item.market_value for item in result.fixed_assets) == result.household.property_market_value
    assert not result.reconciliation_warnings


def test_parser_warns_when_property_or_rent_is_missing():
    payload = make_workbook_bytes(include_household=True)
    book = load_workbook(BytesIO(payload))
    book["BALANCE SHEET"]["C7"] = None
    book["MNTHLY INCOM PLAN"]["E4"] = None
    stream = BytesIO()
    book.save(stream)
    result = parse_workbook(stream.getvalue(), "missing.xlsx")
    codes = {issue.code for issue in result.reconciliation_warnings}
    assert {"missing_property_value", "missing_rent"} <= codes


def test_parser_warns_when_balance_sheet_total_does_not_reconcile():
    payload = make_workbook_bytes(include_household=True)
    book = load_workbook(BytesIO(payload))
    book["BALANCE SHEET"]["C9"] = 83_058_854
    stream = BytesIO()
    book.save(stream)
    result = parse_workbook(stream.getvalue(), "mismatch.xlsx")
    assert any(issue.code == "household_total_mismatch" for issue in result.reconciliation_warnings)


def test_parser_reports_ignored_sheet_without_reading_cells():
    result = parse_workbook(make_workbook_bytes(), "investment.xlsx")
    assert "MF discont." in result.ignored_sheets
    assert "DO_NOT_EXPOSE" not in result.model_dump_json()


def test_parser_assigns_same_source_key_on_repeat_parse():
    payload = make_workbook_bytes()
    first = parse_workbook(payload, "a.xlsx")
    second = parse_workbook(payload, "b.xlsx")
    assert first.transactions[0].source_key == second.transactions[0].source_key


def test_parser_reports_invalid_transaction_date_as_blocking_error():
    result = parse_workbook(make_workbook_bytes(invalid_transaction_date=True), "bad.xlsx")
    issue = next(item for item in result.issues if item.code == "invalid_transaction_date")
    assert issue.severity == "error"
    assert issue.sheet == "Funds XIRR"
    assert issue.row == 2


def test_parser_supports_real_multirow_funds_and_paired_xirr_layout():
    result = parse_workbook(make_real_layout_workbook_bytes(), "investment.xlsx")
    assert result.counts == {"assets": 1, "transactions": 2, "valuations": 1}
    assert [item.amount for item in result.transactions] == [5000, 400000]
    assert result.valuations[0].market_value == 493000
    assert result.assets[0].account_owner == "Nishi"


def test_paired_layout_warns_when_formula_cash_flow_has_no_cached_value():
    result = parse_workbook(
        make_real_layout_workbook_bytes(include_incomplete_cash_flow=True),
        "investment.xlsx",
    )
    issue = next(item for item in result.issues if item.code == "incomplete_cash_flow")
    assert issue.severity == "warning"


def test_paired_layout_recovers_cash_flow_from_date_formatted_amount_cell():
    result = parse_workbook(
        make_real_layout_workbook_bytes(date_formatted_amount=True),
        "investment.xlsx",
    )
    assert result.transactions[0].amount == 5000


def test_paired_layout_warns_when_cash_flow_amount_has_no_date():
    result = parse_workbook(
        make_real_layout_workbook_bytes(include_amount_without_date=True),
        "investment.xlsx",
    )
    issues = [item for item in result.issues if item.code == "incomplete_cash_flow"]
    assert issues[-1].severity == "warning"


def test_parser_extracts_source_ledger_assets_observations_and_property_capital():
    result = parse_workbook(make_source_ledger_workbook_bytes(), "investment.xlsx")

    assert {(asset.owner, asset.name) for asset in result.ledger_assets} == {
        ("Nishi", "Mutual funds"), ("Supriya", "Bank deposits"),
        (None, "Brigade land"), (None, "Gera Office"),
    }
    mutual_fund = next(asset for asset in result.ledger_assets if asset.name == "Mutual funds")
    observations = [item for item in result.asset_observations if item.asset_source_key == mutual_fund.source_key]
    assert [(item.observed_on.isoformat(), item.principal, item.market_value) for item in observations] == [
        ("2024-12-31", None, 1_100_000),
        ("2025-12-31", 1_000_000, 1_250_000),
        ("2026-04-25", 1_100_000, 1_300_000),
    ]
    assert [(flow.occurred_on.isoformat(), flow.amount) for flow in result.ledger_cash_flows] == [
        ("2024-06-01", 500_000), ("2025-06-01", 1_000_000), ("2026-02-01", 500_000),
    ]


def test_parser_preserves_balance_sheet_reporting_lineage_without_importing_totals_as_facts():
    result = parse_workbook(make_source_ledger_workbook_bytes(), "investment.xlsx")

    assert [period.label for period in result.reporting_periods] == ["FY-2024", "FY-2025", "FY-2026"]
    fy_2026 = result.reporting_periods[-1]
    assert {(source.metric, source.source_sheet, source.source_cell) for source in fy_2026.sources} == {
        ("financial_principal", "CURRENT ASSET", "G6"),
        ("financial_market_value", "CURRENT ASSET", "H6"),
        ("property_principal", "FIXED ASSET", "H4"),
        ("property_market_value", "FIXED ASSET", "I4"),
    }
    assert fy_2026.sources[0].observed_on.isoformat() == "2026-04-25"
    assert len(result.asset_observations) == 12
    assert {asset.name for asset in result.ledger_assets}.isdisjoint({"Sub total", "Total", "Notes"})


def test_parser_deduplicates_repeated_columns_for_same_asset_observation():
    result = parse_workbook(
        make_source_ledger_workbook_bytes(duplicate_property_observation=True),
        "investment.xlsx",
    )
    assert len(result.asset_observations) == 12
