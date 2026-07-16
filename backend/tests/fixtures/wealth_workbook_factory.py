from datetime import date
from io import BytesIO

from openpyxl import Workbook


def make_workbook_bytes(
    *, invalid_transaction_date: bool = False, include_household: bool = False,
    duplicate_fund_row: bool = False, duplicate_transaction_row: bool = False,
) -> bytes:
    workbook = Workbook()
    funds = workbook.active
    funds.title = "FUNDS"
    funds.append(["Fund", "Principal", "Market Value", "Category"])
    funds.append(["Example Mid Cap", 400000, 520000, "Mid cap"])
    if duplicate_fund_row:
        funds.append(["Example Mid Cap", 10000, 12000, "Mid cap"])

    xirr = workbook.create_sheet("Funds XIRR")
    xirr.append(["Fund", "Date", "Amount", "Units", "NAV"])
    xirr.append([
        "Example Mid Cap",
        "not-a-date" if invalid_transaction_date else date(2025, 3, 6),
        400000,
        4025.481297,
        99.367,
    ])
    if duplicate_transaction_row:
        xirr.append(["Example Mid Cap", date(2025, 3, 6), 400000, 4025.481297, 99.367])

    balance = workbook.create_sheet("BALANCE SHEET")
    if include_household:
        balance.append(["BALANCE SHEET"])
        balance.append([])
        balance.append(["ASSET TYPE", "FY-2025", "FY-2026", "FY-2027"])
        balance.append(["Current assests principal (year end)", 39_892_239.25, 40_953_839.25, None])
        balance.append(["Current assets market value (year end)", 44_110_481.25, 44_658_852.25, None])
        balance.append(["Fixed assests principal (year end)", 17_709_216, 17_709_216, None])
        balance.append(["Fixed assests market value (year end)", 38_400_000, 38_400_000, None])
        balance.append(["TOTAL ASSETS PRINCIPAL", 57_601_455.25, 58_663_055.25, None])
        balance.append(["TOTAL ASSETS MARKET VALUE", 82_510_481.25, 83_058_852.25, None])

    fixed = workbook.create_sheet("FIXED ASSET")
    if include_household:
        fixed.append(["S No.", "Name", "Desc", "01/9/2025\nPRINCIPAL", "01/9/2025\nMKT VALUE"])
        fixed.append([1, "Brigade land", "Residential Land", 9_959_216, 21_600_000])
        fixed.append([2, "Amrapali Flat G Noida M8-306", "Residential flat", 2_050_000, 6_800_000])
        fixed.append([3, "Gera Office Pune 733", "Office", 5_700_000, 10_000_000])
        fixed.append([None, None, "Total", 17_709_216, 38_400_000])

    income = workbook.create_sheet("MNTHLY INCOM PLAN")
    if include_household:
        income.append([None, None, "Option 1"])
        income.append(["Desc", "Principal Amount", None, "Yearly Income", "Monthly Income"])
        income.append(["Gera office rent", None, None, 360_000, 30_000])
        income.append(["Golfhome rent", None, None, 168_000, 14_000])

    for sheet_name in ("CURRENT ASSET", "Final XIRR", "EQUITY", "GOALS", "Gera office roi"):
        workbook.create_sheet(sheet_name)

    ignored = workbook.create_sheet("MF discont.")
    ignored.append(["DO_NOT_EXPOSE"])
    workbook.create_sheet("Property Cal.")
    workbook.create_sheet("REMIT")
    workbook.create_sheet("STOCKS RECMDN")

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def make_real_layout_workbook_bytes(
    *, include_incomplete_cash_flow: bool = False, date_formatted_amount: bool = False,
    include_amount_without_date: bool = False,
) -> bytes:
    workbook = Workbook()
    funds = workbook.active
    funds.title = "FUNDS"
    funds.append([None, None, None, "Last Updated", date(2026, 4, 25)])
    funds.append(["S NO.", "NAME", "MODE", "FUND NAME", "CATEGORY", "DATE STARTED", "LAST UPDATED", "PRINCIPAL", "MARKET VALUE"])
    funds.append([1, "Nishi", "Lumpsum", "Example Mid Cap", "MID_CAP", None, date(2025, 3, 6), 405000, 493000])
    funds.append([None, "Nishi", "Lumpsum", "SUB TOTAL", None, None, None, 405000, 493000])

    xirr = workbook.create_sheet("Funds XIRR")
    xirr.append(["Nishi", "MID_CAP"])
    xirr.append(["Example Mid Cap", None])
    xirr.append(["XIRR", 0.18])
    xirr.append([date(2025, 1, 27), -5000])
    if date_formatted_amount:
        xirr.cell(row=4, column=2).number_format = "dd-mmm-yyyy"
    xirr.append([date(2025, 3, 6), -400000])
    xirr.append([date(2026, 4, 25), 493000])
    if include_incomplete_cash_flow:
        xirr.append([date(2025, 4, 1), None])
    if include_amount_without_date:
        xirr.append([None, -10000])

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def make_source_ledger_workbook_bytes(*, duplicate_property_observation: bool = False) -> bytes:
    workbook = Workbook()
    current = workbook.active
    current.title = "CURRENT ASSET"
    current.append([None, None, None, date(2024, 12, 31), date(2025, 12, 31), date(2025, 12, 31), date(2026, 4, 25), date(2026, 4, 25)])
    current.append(["OWNER", "DESCRIPTION", "CATEGORY", "MKT VALUE", "PRINCIPAL", "MKT VALUE", "PRINCIPAL", "MKT VALUE"])
    current.append(["Nishi", "Mutual funds", "Equity", 1_100_000, 1_000_000, 1_250_000, 1_100_000, 1_300_000])
    current.append(["Supriya", "Bank deposits", "Debt", 500_000, 500_000, 525_000, 510_000, 540_000])
    current.append(["Sub total", None, None, 1_600_000, 1_500_000, 1_775_000, 1_610_000, 1_840_000])
    current.append(["Total", None, None, 1_600_000, 1_500_000, 1_775_000, 1_610_000, 1_840_000])
    current.append(["Notes", "not a transaction", None, None, None, None, None, None])

    fixed = workbook.create_sheet("FIXED ASSET")
    fixed.append(["S No.", "Name", "Desc", "31/12/2024\nPRINCIPAL", "31/12/2024\nMKT VALUE", "31/12/2025\nPRINCIPAL", "31/12/2025\nMKT VALUE", "25/04/2026\nPRINCIPAL", "25/04/2026\nMKT VALUE"])
    fixed.append([1, "Brigade land", "Residential land", 8_000_000, 16_000_000, 9_000_000, 18_000_000, 9_500_000, 19_000_000])
    fixed.append([2, "Gera Office", "Office", 5_000_000, 8_000_000, 5_000_000, 9_000_000, 5_000_000, 10_000_000])
    fixed.append([None, None, "Total", "=SUM(D2:D3)", "=SUM(E2:E3)", "=SUM(F2:F3)", "=SUM(G2:G3)", "=SUM(H2:H3)", "=SUM(I2:I3)"])
    if duplicate_property_observation:
        fixed["J1"] = "25/04/2026\nPRINCIPAL"
        fixed["K1"] = "25/04/2026\nMKT VALUE"
        fixed["J2"], fixed["K2"] = 9_500_000, 19_000_000
        fixed["J3"], fixed["K3"] = 5_000_000, 10_000_000
    fixed.append([])
    fixed.append(["Brigade payments"])
    fixed.append([None, "Date", "Amount", "Date", "Amount", None, "Date", "Amount"])
    fixed.append([None, date(2024, 6, 1), 500_000, date(2025, 6, 1), 1_000_000, None, date(2026, 2, 1), 500_000])
    fixed.append([None, "Total 2024", 500_000, "Total 2025", 1_000_000, None, "Total 2026", 500_000])
    fixed.append([None, date(2026, 5, 1), 99_000_000])

    balance = workbook.create_sheet("BALANCE SHEET")
    balance.append(["ASSET TYPE", "FY-2024", "FY-2025", "FY-2026"])
    balance.append(["Current assests principal (year end)", "='CURRENT ASSET'!E6", "='CURRENT ASSET'!E6", "='CURRENT ASSET'!G6"])
    balance.append(["Current assets market value (year end)", "='CURRENT ASSET'!D6", "='CURRENT ASSET'!F6", "='CURRENT ASSET'!H6"])
    balance.append(["Fixed assests principal (year end)", "='FIXED ASSET'!D4", "='FIXED ASSET'!F4", "='FIXED ASSET'!H4"])
    balance.append(["Fixed assests market value (year end)", "='FIXED ASSET'!E4", "='FIXED ASSET'!G4", "='FIXED ASSET'!I4"])
    balance.append(["TOTAL ASSETS PRINCIPAL", "=B2+B4", "=C2+C4", "=D2+D4"])
    balance.append(["TOTAL ASSETS MARKET VALUE", "=B3+B5", "=C3+C5", "=D3+D5"])

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()
